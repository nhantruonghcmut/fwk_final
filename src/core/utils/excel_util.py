"""
Excel utility for reading and writing Excel files using openpyxl.
"""
import os
import time
from datetime import datetime
from typing import Any, Dict, List, Optional, Union
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from src.core.utils.report_logger import ReportLogger


class ExcelUtil:
    """Utility class for Excel operations using openpyxl."""
    
    def __init__(self):
        self.logger = ReportLogger()
        
    def create_workbook(self) -> Workbook:
        """Create new workbook."""
        try:
            workbook = Workbook()
            self.logger.info("Created new Excel workbook")
            return workbook
        except Exception as e:
            self.logger.log_error(e, "create_workbook")
            raise
            
    def load_workbook(self, file_path: str) -> Workbook:
        """Load existing workbook."""
        try:
            workbook = load_workbook(file_path)
            self.logger.info(f"Loaded Excel workbook: {file_path}")
            return workbook
        except Exception as e:
            self.logger.log_error(e, "load_workbook")
            raise
            
    def save_workbook(self, workbook: Workbook, file_path: str):
        """Save workbook to file."""
        try:
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            workbook.save(file_path)
            self.logger.info(f"Saved Excel workbook: {file_path}")
        except Exception as e:
            self.logger.log_error(e, "save_workbook")
            raise
            
    def write_data_to_sheet(self, workbook: Workbook, sheet_name: str, data: List[List[Any]], 
                           headers: List[str] = None, start_row: int = 1):
        """Write data to worksheet."""
        try:
            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.create_sheet(sheet_name)
                
            # Write headers if provided
            if headers:
                for col, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=start_row, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    
            # Write data
            for row_idx, row_data in enumerate(data, start_row + (1 if headers else 0)):
                for col_idx, value in enumerate(row_data, 1):
                    worksheet.cell(row=row_idx, column=col_idx, value=value)
                    
            # Auto-adjust column widths
            self._auto_adjust_columns(worksheet)
            
            self.logger.info(f"Written {len(data)} rows to sheet '{sheet_name}'")
            
        except Exception as e:
            self.logger.log_error(e, "write_data_to_sheet")
            raise
            
    def read_data_from_sheet(self, workbook: Workbook, sheet_name: str, 
                           start_row: int = 1, end_row: int = None) -> List[List[Any]]:
        """Read data from worksheet."""
        try:
            worksheet = workbook[sheet_name]
            data = []
            
            max_row = end_row if end_row else worksheet.max_row
            
            for row in range(start_row, max_row + 1):
                row_data = []
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    row_data.append(cell_value)
                data.append(row_data)
                
            self.logger.info(f"Read {len(data)} rows from sheet '{sheet_name}'")
            return data
            
        except Exception as e:
            self.logger.log_error(e, "read_data_from_sheet")
            raise
            
    def read_data_as_dict(self, workbook: Workbook, sheet_name: str, 
                         start_row: int = 1, end_row: int = None) -> List[Dict[str, Any]]:
        """Read data as list of dictionaries."""
        try:
            worksheet = workbook[sheet_name]
            data = []
            
            # Get headers from first row
            headers = []
            for col in range(1, worksheet.max_column + 1):
                header = worksheet.cell(row=start_row, column=col).value
                headers.append(header or f"Column_{col}")
                
            max_row = end_row if end_row else worksheet.max_row
            
            # Read data rows
            for row in range(start_row + 1, max_row + 1):
                row_dict = {}
                for col, header in enumerate(headers, 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    row_dict[header] = cell_value
                data.append(row_dict)
                
            self.logger.info(f"Read {len(data)} rows as dictionaries from sheet '{sheet_name}'")
            return data
            
        except Exception as e:
            self.logger.log_error(e, "read_data_as_dict")
            raise
            
    def create_test_report(self, test_results: List[Dict[str, Any]], suite_name: str) -> str:
        """Create test execution report."""
        try:
            workbook = self.create_workbook()
            
            # Test Summary Sheet
            self._create_test_summary_sheet(workbook, test_results, suite_name)
            
            # Test Details Sheet
            self._create_test_details_sheet(workbook, test_results)
            
            # Test Statistics Sheet
            self._create_test_statistics_sheet(workbook, test_results)
            
            # Save workbook
            file_path = f"reports/excel/test_report_{suite_name}_{int(time.time())}.xlsx"
            self.save_workbook(workbook, file_path)
            
            return file_path
            
        except Exception as e:
            self.logger.log_error(e, "create_test_report")
            raise
            
    def _create_test_summary_sheet(self, workbook: Workbook, test_results: List[Dict[str, Any]], suite_name: str):
        """Create test summary sheet."""
        try:
            worksheet = workbook.active
            worksheet.title = "Test Summary"
            
            # Summary data
            total_tests = len(test_results)
            passed_tests = len([t for t in test_results if t.get("result") == "PASSED"])
            failed_tests = len([t for t in test_results if t.get("result") == "FAILED"])
            skipped_tests = len([t for t in test_results if t.get("result") == "SKIPPED"])
            pass_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
            
            summary_data = [
                ["Test Suite", suite_name],
                ["Total Tests", total_tests],
                ["Passed", passed_tests],
                ["Failed", failed_tests],
                ["Skipped", skipped_tests],
                ["Pass Rate", f"{pass_rate:.1f}%"],
                ["Execution Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
            ]
            
            self.write_data_to_sheet(workbook, "Test Summary", summary_data, 
                                   headers=["Metric", "Value"], start_row=1)
            
        except Exception as e:
            self.logger.log_error(e, "_create_test_summary_sheet")
            raise
            
    def _create_test_details_sheet(self, workbook: Workbook, test_results: List[Dict[str, Any]]):
        """Create test details sheet."""
        try:
            worksheet = workbook.create_sheet("Test Details")
            
            # Headers
            headers = ["Test Name", "Test File", "Test Class", "Test Method", 
                      "Result", "Duration (s)", "Error Message", "Screenshots"]
            
            # Data
            data = []
            for result in test_results:
                row = [
                    result.get("name", ""),
                    result.get("file", ""),
                    result.get("class", ""),
                    result.get("method", ""),
                    result.get("result", ""),
                    result.get("duration", 0),
                    result.get("error", ""),
                    result.get("screenshots", "")
                ]
                data.append(row)
                
            self.write_data_to_sheet(workbook, "Test Details", data, headers=headers)
            
        except Exception as e:
            self.logger.log_error(e, "_create_test_details_sheet")
            raise
            
    def _create_test_statistics_sheet(self, workbook: Workbook, test_results: List[Dict[str, Any]]):
        """Create test statistics sheet."""
        try:
            worksheet = workbook.create_sheet("Test Statistics")
            
            # Calculate statistics
            durations = [r.get("duration", 0) for r in test_results if r.get("duration")]
            avg_duration = sum(durations) / len(durations) if durations else 0
            min_duration = min(durations) if durations else 0
            max_duration = max(durations) if durations else 0
            
            # Group by result
            result_counts = {}
            for result in test_results:
                result_type = result.get("result", "UNKNOWN")
                result_counts[result_type] = result_counts.get(result_type, 0) + 1
                
            # Group by test file
            file_counts = {}
            for result in test_results:
                test_file = result.get("file", "Unknown")
                file_counts[test_file] = file_counts.get(test_file, 0) + 1
                
            # Statistics data
            stats_data = [
                ["Average Duration", f"{avg_duration:.2f}s"],
                ["Minimum Duration", f"{min_duration:.2f}s"],
                ["Maximum Duration", f"{max_duration:.2f}s"],
                ["", ""],
                ["Result Distribution", ""],
                *[[result, count] for result, count in result_counts.items()],
                ["", ""],
                ["Tests by File", ""],
                *[[file, count] for file, count in file_counts.items()]
            ]
            
            self.write_data_to_sheet(workbook, "Test Statistics", stats_data, 
                                   headers=["Metric", "Value"], start_row=1)
            
        except Exception as e:
            self.logger.log_error(e, "_create_test_statistics_sheet")
            raise
            
    def _auto_adjust_columns(self, worksheet):
        """Auto-adjust column widths."""
        try:
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                        
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            self.logger.log_error(e, "_auto_adjust_columns")
            
    def create_data_driven_sheet(self, workbook: Workbook, sheet_name: str, 
                                test_data: List[Dict[str, Any]]) -> str:
        """Create data-driven test sheet."""
        try:
            if not test_data:
                return ""
                
            # Get headers from first data item
            headers = list(test_data[0].keys())
            
            # Convert data to list of lists
            data = []
            for item in test_data:
                row = [item.get(header, "") for header in headers]
                data.append(row)
                
            self.write_data_to_sheet(workbook, sheet_name, data, headers=headers)
            
            self.logger.info(f"Created data-driven sheet '{sheet_name}' with {len(data)} rows")
            return sheet_name
            
        except Exception as e:
            self.logger.log_error(e, "create_data_driven_sheet")
            raise
            
    def read_test_data(self, file_path: str, sheet_name: str = None) -> List[Dict[str, Any]]:
        """Read test data from Excel file."""
        try:
            workbook = self.load_workbook(file_path)
            
            if sheet_name:
                return self.read_data_as_dict(workbook, sheet_name)
            else:
                # Read from first sheet
                first_sheet = workbook.sheetnames[0]
                return self.read_data_as_dict(workbook, first_sheet)
                
        except Exception as e:
            self.logger.log_error(e, "read_test_data")
            raise
            
    def update_cell(self, workbook: Workbook, sheet_name: str, row: int, col: int, value: Any):
        """Update specific cell."""
        try:
            worksheet = workbook[sheet_name]
            worksheet.cell(row=row, column=col, value=value)
            self.logger.debug(f"Updated cell {sheet_name}!{row}:{col} = {value}")
        except Exception as e:
            self.logger.log_error(e, "update_cell")
            raise
            
    def get_cell_value(self, workbook: Workbook, sheet_name: str, row: int, col: int) -> Any:
        """Get cell value."""
        try:
            worksheet = workbook[sheet_name]
            return worksheet.cell(row=row, column=col).value
        except Exception as e:
            self.logger.log_error(e, "get_cell_value")
            raise
            
    def add_chart(self, workbook: Workbook, sheet_name: str, chart_type: str, 
                  data_range: str, position: str = "E2"):
        """Add chart to worksheet."""
        try:
            from openpyxl.chart import BarChart, LineChart, PieChart
            
            worksheet = workbook[sheet_name]
            
            if chart_type.lower() == "bar":
                chart = BarChart()
            elif chart_type.lower() == "line":
                chart = LineChart()
            elif chart_type.lower() == "pie":
                chart = PieChart()
            else:
                chart = BarChart()
                
            chart.title = f"{chart_type.title()} Chart"
            chart.add_data(worksheet[data_range])
            worksheet.add_chart(chart, position)
            
            self.logger.info(f"Added {chart_type} chart to sheet '{sheet_name}'")
            
        except Exception as e:
            self.logger.log_error(e, "add_chart")
            raise
            
    def format_cells(self, workbook: Workbook, sheet_name: str, 
                    cell_range: str, font: Font = None, fill: PatternFill = None,
                    alignment: Alignment = None, border: Border = None):
        """Format cells in worksheet."""
        try:
            worksheet = workbook[sheet_name]
            
            for row in worksheet[cell_range]:
                for cell in row:
                    if font:
                        cell.font = font
                    if fill:
                        cell.fill = fill
                    if alignment:
                        cell.alignment = alignment
                    if border:
                        cell.border = border
                        
            self.logger.info(f"Formatted cells {cell_range} in sheet '{sheet_name}'")
            
        except Exception as e:
            self.logger.log_error(e, "format_cells")
            raise
            
    def merge_cells(self, workbook: Workbook, sheet_name: str, cell_range: str):
        """Merge cells in worksheet."""
        try:
            worksheet = workbook[sheet_name]
            worksheet.merge_cells(cell_range)
            self.logger.info(f"Merged cells {cell_range} in sheet '{sheet_name}'")
        except Exception as e:
            self.logger.log_error(e, "merge_cells")
            raise
            
    def get_sheet_names(self, workbook: Workbook) -> List[str]:
        """Get all sheet names."""
        try:
            return workbook.sheetnames
        except Exception as e:
            self.logger.log_error(e, "get_sheet_names")
            return []
            
    def delete_sheet(self, workbook: Workbook, sheet_name: str):
        """Delete worksheet."""
        try:
            workbook.remove(workbook[sheet_name])
            self.logger.info(f"Deleted sheet '{sheet_name}'")
        except Exception as e:
            self.logger.log_error(e, "delete_sheet")
            raise
            
    def copy_sheet(self, workbook: Workbook, source_sheet: str, target_sheet: str):
        """Copy worksheet."""
        try:
            source = workbook[source_sheet]
            target = workbook.copy_worksheet(source)
            target.title = target_sheet
            self.logger.info(f"Copied sheet '{source_sheet}' to '{target_sheet}'")
        except Exception as e:
            self.logger.log_error(e, "copy_sheet")
            raise
            
    def get_workbook_info(self, workbook: Workbook) -> Dict[str, Any]:
        """Get workbook information."""
        try:
            return {
                "sheet_names": workbook.sheetnames,
                "sheet_count": len(workbook.sheetnames),
                "active_sheet": workbook.active.title
            }
        except Exception as e:
            self.logger.log_error(e, "get_workbook_info")
            return {}
            
    def generate_allure_style_report(self, suites_data: List[Dict[str, Any]], 
                                    report_name: str = "Test Execution Report") -> str:
        """
        Generate Allure-style Excel report với sheet tổng quan và chi tiết từng suite/testcase/steps.
        
        Args:
            suites_data: List of suite data, mỗi suite có format:
                {
                    "name": "suite_name",
                    "passed": 10,
                    "failed": 2,
                    "skipped": 1,
                    "total": 13,
                    "duration": 120.5,
                    "test_cases": [
                        {
                            "name": "test_name",
                            "file": "test_file.py",
                            "class": "TestClass",
                            "method": "test_method",
                            "result": "PASSED",
                            "duration": 5.2,
                            "error": None,
                            "steps": [
                                {
                                    "name": "step_name",
                                    "data": {...},
                                    "timestamp": "...",
                                    "status": "PASSED"
                                }
                            ],
                            "screenshots": [...]
                        }
                    ]
                }
            report_name: Tên báo cáo
            
        Returns:
            Đường dẫn file Excel đã tạo
        """
        try:
            workbook = self.create_workbook()
            
            # Sheet 1: Overview
            self._create_overview_sheet(workbook, suites_data, report_name)
            
            # Sheet 2+: Details for each suite
            # for suite in suites_data:
            #     self._create_suite_detail_sheet(workbook, suite)
            
            # Save workbook
            file_path = f"reports/excel/allure_style_report_{int(time.time())}.xlsx"
            self.save_workbook(workbook, file_path)
            
            self.logger.info(f"Generated Allure-style Excel report: {file_path}")
            return file_path
            
        except Exception as e:
            self.logger.log_error(e, "generate_allure_style_report")
            raise
            
    def _create_overview_sheet(self, workbook: Workbook, suites_data: List[Dict[str, Any]], 
                               report_name: str):
        """Create overview sheet with statistics for all suites."""
        try:
            worksheet = workbook.active
            worksheet.title = "Overview"
            
            # Title
            worksheet.merge_cells('A1:G1')
            title_cell = worksheet['A1']
            title_cell.value = report_name
            title_cell.font = Font(bold=True, size=16)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            worksheet.row_dimensions[1].height = 25
            
            # Tổng hợp tất cả suites
            total_passed = sum(s.get("passed", 0) for s in suites_data)
            total_failed = sum(s.get("failed", 0) for s in suites_data)
            total_skipped = sum(s.get("skipped", 0) for s in suites_data)
            total_tests = total_passed + total_failed + total_skipped
            total_duration = sum(s.get("duration", 0) for s in suites_data)
            pass_rate = (total_passed / total_tests * 100) if total_tests > 0 else 0
            
            # Summary section
            row = 3
            worksheet.cell(row=row, column=1, value="Overview")
            worksheet.cell(row=row, column=1).font = Font(bold=True, size=14)
            row += 1
            
            summary_headers = ["Metric", "Value"]
            summary_data = [
                ["Total Tests", total_tests],
                ["Passed", total_passed],
                ["Failed", total_failed],
                ["Skipped", total_skipped],
                ["Pass Rate", f"{pass_rate:.1f}%"],
                ["Total Duration", f"{total_duration:.2f} seconds"],
                ["Execution Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
            ]
            
            # Write summary headers
            for col, header in enumerate(summary_headers, 1):
                cell = worksheet.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            row += 1
            
            # Write summary data
            for data_row in summary_data:
                for col, value in enumerate(data_row, 1):
                    worksheet.cell(row=row, column=col, value=value)
                row += 1
            
            row += 2
            
            # Suite summary table
            worksheet.cell(row=row, column=1, value="Test Suite Statistics")
            worksheet.cell(row=row, column=1).font = Font(bold=True, size=14)
            row += 1
            
            suite_headers = ["Test Suite", "Total", "Passed", "Failed", "Skipped", "Pass Rate", "Duration (s)"]
            
            # Write suite headers
            for col, header in enumerate(suite_headers, 1):
                cell = worksheet.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            row += 1
            
            # Write suite data
            for suite in suites_data:
                suite_total = suite.get("total", 0)
                suite_passed = suite.get("passed", 0)
                suite_failed = suite.get("failed", 0)
                suite_skipped = suite.get("skipped", 0)
                suite_pass_rate = (suite_passed / suite_total * 100) if suite_total > 0 else 0
                suite_duration = suite.get("duration", 0)
                
                suite_row = [
                    suite.get("name", "Unknown"),
                    suite_total,
                    suite_passed,
                    suite_failed,
                    suite_skipped,
                    f"{suite_pass_rate:.1f}%",
                    f"{suite_duration:.2f}"
                ]
                
                for col, value in enumerate(suite_row, 1):
                    cell = worksheet.cell(row=row, column=col, value=value)
                    # Color code based on pass rate
                    if suite_pass_rate == 100:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif suite_pass_rate < 50:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                row += 1
            
            # Auto-adjust columns
            self._auto_adjust_columns(worksheet)
            
        except Exception as e:
            self.logger.log_error(e, "_create_overview_sheet")
            raise
            
    def _create_suite_detail_sheet(self, workbook: Workbook, suite: Dict[str, Any]):
        """Create detail sheet for a test suite."""
        try:
            suite_name = suite.get("name", "Unknown Suite")
            # Excel sheet name has 31 character limit
            sheet_name = suite_name[:31] if len(suite_name) <= 31 else suite_name[:28] + "..."
            worksheet = workbook.create_sheet(sheet_name)
            
            # Title
            worksheet.merge_cells('A1:F1')
            title_cell = worksheet['A1']
            title_cell.value = f"Test Suite: {suite_name}"
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            worksheet.row_dimensions[1].height = 20
            
            # Suite summary
            row = 3
            worksheet.cell(row=row, column=1, value="Suite Information")
            worksheet.cell(row=row, column=1).font = Font(bold=True, size=12)
            row += 1
            
            suite_total = suite.get("total", 0)
            suite_passed = suite.get("passed", 0)
            suite_failed = suite.get("failed", 0)
            suite_skipped = suite.get("skipped", 0)
            suite_pass_rate = (suite_passed / suite_total * 100) if suite_total > 0 else 0
            suite_duration = suite.get("duration", 0)
            
            suite_info_headers = ["Metric", "Value"]
            suite_info_data = [
                ["Total Tests", suite_total],
                ["Passed", suite_passed],
                ["Failed", suite_failed],
                ["Skipped", suite_skipped],
                ["Pass Rate", f"{suite_pass_rate:.1f}%"],
                ["Execution Duration", f"{suite_duration:.2f} seconds"]
            ]
            
            # Write suite info headers
            for col, header in enumerate(suite_info_headers, 1):
                cell = worksheet.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            row += 1
            
            # Write suite info data
            for data_row in suite_info_data:
                for col, value in enumerate(data_row, 1):
                    worksheet.cell(row=row, column=col, value=value)
                row += 1
            
            row += 2
            
            # Test cases list
            test_cases = suite.get("test_cases", [])
            if test_cases:
                worksheet.cell(row=row, column=1, value="Test Cases List")
                worksheet.cell(row=row, column=1).font = Font(bold=True, size=12)
                row += 1
                
                # Test cases summary table
                test_case_headers = ["Test Case", "Result", "Duration (s)", "File", "Class", "Method"]
                
                # Write test case headers
                for col, header in enumerate(test_case_headers, 1):
                    cell = worksheet.cell(row=row, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                row += 1
                
                # Write test case data
                for test_case in test_cases:
                    test_case_row = [
                        test_case.get("name", "Unknown"),
                        test_case.get("result", "UNKNOWN"),
                        test_case.get("duration", 0),
                        test_case.get("file", ""),
                        test_case.get("class", "N/A"),
                        test_case.get("method", "N/A")
                    ]
                    
                    for col, value in enumerate(test_case_row, 1):
                        cell = worksheet.cell(row=row, column=col, value=value)
                        # Color code based on result
                        result = test_case.get("result", "")
                        if result == "PASSED":
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif result == "FAILED":
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        elif result == "SKIPPED":
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    
                    row += 1
                
                row += 2
                
                # Details for each test case with steps
                for test_case in test_cases:
                    row = self._add_test_case_details(worksheet, test_case, row)
                    row += 2
            
            # Auto-adjust columns
            self._auto_adjust_columns(worksheet)
            
        except Exception as e:
            self.logger.log_error(e, "_create_suite_detail_sheet")
            raise
            
    def _add_test_case_details(self, worksheet, test_case: Dict[str, Any], start_row: int) -> int:
        """Add test case details with steps to worksheet. Returns next row."""
        try:
            row = start_row
            
            test_name = test_case.get("name", "Unknown Test")
            test_result = test_case.get("result", "UNKNOWN")
            
            # Test case header
            worksheet.merge_cells(f'A{row}:F{row}')
            header_cell = worksheet.cell(row=row, column=1, value=f"Test Case: {test_name} [{test_result}]")
            header_cell.font = Font(bold=True, size=11)
            
            # Color code header
            if test_result == "PASSED":
                header_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif test_result == "FAILED":
                header_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif test_result == "SKIPPED":
                header_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            
            row += 1
            
            # Test case info
            test_info = [
                ["File", test_case.get("file", "Unknown")],
                ["Class", test_case.get("class", "N/A")],
                ["Method", test_case.get("method", "N/A")],
                ["Result", test_result],
                ["Duration", f"{test_case.get('duration', 0):.2f} seconds"]
            ]
            
            for info in test_info:
                worksheet.cell(row=row, column=1, value=info[0]).font = Font(bold=True)
                worksheet.cell(row=row, column=2, value=info[1])
                row += 1
            
            row += 1
            
            # Test steps
            steps = test_case.get("steps", [])
            if steps:
                worksheet.cell(row=row, column=1, value="Test Steps")
                worksheet.cell(row=row, column=1).font = Font(bold=True, size=10)
                row += 1
                
                # Step headers
                step_headers = ["#", "Step Name", "Status", "Timestamp", "Data"]
                for col, header in enumerate(step_headers, 1):
                    cell = worksheet.cell(row=row, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                
                row += 1
                
                # Step data
                for idx, step in enumerate(steps, 1):
                    step_name = step.get("name", "Unknown Step")
                    step_status = step.get("status", "UNKNOWN")
                    step_timestamp = step.get("timestamp", "")
                    step_data = step.get("data", "")
                    
                    step_row = [
                        idx,
                        step_name,
                        step_status,
                        step_timestamp,
                        str(step_data)[:100] if step_data else ""  # Truncate long data
                    ]
                    
                    for col, value in enumerate(step_row, 1):
                        cell = worksheet.cell(row=row, column=col, value=value)
                        # Color code based on step status
                        if step_status == "PASSED":
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif step_status == "FAILED":
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    
                    row += 1
                
                row += 1
            
            # Error information nếu có
            error = test_case.get("error")
            if error:
                worksheet.cell(row=row, column=1, value="Error Details")
                worksheet.cell(row=row, column=1).font = Font(bold=True, color="FF0000", size=10)
                row += 1
                
                error_lines = str(error).split('\n')
                for line in error_lines[:10]:  # Limit error lines
                    worksheet.cell(row=row, column=1, value=line[:200])  # Truncate long lines
                    worksheet.cell(row=row, column=1).font = Font(name="Courier", size=8)
                    row += 1
                
                row += 1
            
            # Screenshots if available
            screenshots = test_case.get("screenshots", [])
            if screenshots:
                worksheet.cell(row=row, column=1, value="Screenshots")
                worksheet.cell(row=row, column=1).font = Font(bold=True, size=10)
                row += 1
                
                for screenshot in screenshots[:5]:  # Limit số lượng screenshots
                    if isinstance(screenshot, str):
                        worksheet.cell(row=row, column=1, value=screenshot)
                        row += 1
            
            return row
            
        except Exception as e:
            self.logger.log_error(e, "_add_test_case_details")
            return start_row